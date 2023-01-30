from tkinter import *
from openpyxl import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from openpyxl.styles import Font
from tkcalendar import *
import time

root = Tk()
root.title('Customers List - Name')
root.resizable(width=FALSE, height=FALSE)
root.geometry("800x550")

q= StringVar()
name= StringVar()
store= StringVar()
status = StringVar()
date= StringVar()
coment= StringVar()
title= StringVar()
#Fecha y hora
def clockdate():
    hour= time.strftime("%I")
    minute= time.strftime("%M")
    day= time.strftime("%d")
    month= time.strftime("%m")
    year= time.strftime("%Y")
    am_pm= time.strftime("%p")
    hoylabel.config(text="Hoy es: " + day + "/" + month + "/" + year)
    horalabel.config(text="Hora: " + hour + ":" + minute + " " + am_pm )
    horalabel.after(1000, clockdate)

#BACK END
#FUNCTIONS

#Function Creates New File and Saves its in the PC
def newfile():
    global count
    cleartreeview()
    count= 0
    global file
    file = filedialog.asksaveasfilename(title='Guardando dirección del Archivo', defaultextension='*.xlsx', filetypes=(('Excel File', '*.xlsx'),))
    global wb
    global ws
    wb = Workbook()
    ws = wb.active
    headers = ['Hotel/Casa de Buceo', 'Fecha', 'Actividad ó Paciente', 'Observaciones/Servicio',
               'Notas']
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    for i in range(1, 6):
        ws.cell(row=1, column=i).value = headers[i - 1]
        ws.cell(row=1, column=i).font = Font(bold=True)
    wb.save(file)
    totallabel.config(text="Total Clientes: " + str(count))
    statbar.config(text=file)

#Function Opens Existing File in the PC
def openfile():
    cleartreeview()
    global file
    global wb1
    global ws1
    global count
    file = filedialog.askopenfilename(title='Abriendo el Archivo', defaultextension='*.xlsx',
                                                filetypes=(('Excel Files', '*xlsx'),))
    wb1 = load_workbook(file)
    ws1 = wb1.active
    headers = ['Hotel/Casa de Buceo', 'Fecha', 'Actividad ó Paciente', 'Observaciones/Servicio',
                   'Notas']
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 40
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 40
    for i in range(1, 6):
        ws1.cell(row=1, column=i).value = headers[i - 1]
        ws1.cell(row=1, column=i).font = Font(bold=True)
    max_row_num = ws1.max_row
    count= 1
    for i in range(2, max_row_num + 1):
        cell_obj= ws1.cell(row=i, column=1)
        cell_obj1 = ws1.cell(row=i, column=2)
        cell_obj2 = ws1.cell(row=i, column=3)
        cell_obj3 = ws1.cell(row=i, column=4)
        cell_obj4 = ws1.cell(row=i, column=5)
        viewer.insert('', END, iid=count, values=(count, cell_obj.value, cell_obj1.value, cell_obj2.value, cell_obj3.value, cell_obj4.value))
        count +=1
    totallabel.config(text="Total Clientes: " + str(count-1))
    statbar.config(text=file)

#Function Inputs data into de Excele WS & in the Treeview
def input():
    global file
    global count
    wb2= load_workbook(filename=file)
    ws2= wb2.active
    global current_row
    current_row= ws2.max_row
    ws2.cell(row=current_row + 1, column=1).value = nmstore_entry.get()
    ws2.cell(row=current_row + 1, column=2).value = actentry.get()
    ws2.cell(row=current_row + 1, column=3).value = status.get()
    ws2.cell(row=current_row + 1, column=4).value = date_entry.get()
    ws2.cell(row=current_row + 1, column=5).value = cmnt_entry.get()
    # guarda el registro en el archivo
    wb2.save(file)
    count += 1
    aviso_guardado = messagebox.showinfo(message="Tienda Guardada", title='Aviso')
    cell_obj = ws2.cell(row=current_row + 1, column=1)
    cell_obj1 = ws2.cell(row=current_row + 1, column=2)
    cell_obj2 = ws2.cell(row=current_row + 1, column=3)
    cell_obj3 = ws2.cell(row=current_row + 1, column=4)
    cell_obj4 = ws2.cell(row=current_row + 1, column=5)
    viewer.insert('', END, iid= count+1, values=(count, cell_obj.value, cell_obj1.value, cell_obj2.value, cell_obj3.value,cell_obj4.value))

    # limpia despues de guardar el resgito
    clearinput()
    totallabel.config(text="Total Clientes: " + str(count-1))
#Function to clear the Input Entry's
def clearinput():
    nmstore_entry.delete(0, END)
    actentry.delete(0, END)
    fechaentry.delete(0, END)
    date_entry.delete(0, END)
    cmnt_entry.delete(0, END)

#Function to Clear TreeView
def cleartreeview():
    remove= viewer.get_children()
    for items in remove:
        viewer.delete(items)

def getrowsel(event):
    global item
    rowid= viewer.identify_row(event.y)
    item= viewer.item(viewer.focus())
    store.set(item['values'][1])
    name.set(item['values'][3])
    status.set(item['values'][2])
    date.set(item['values'][4])
    coment.set(item['values'][5])

def delete():
    if messagebox.askyesno("Confirma Delete", "Estas seguro de borrar esta cliente?"):
        global file

        id= int(viewer.selection()[0])
        wb= load_workbook(filename=file)
        ws= wb.active
        ws.delete_rows(id)
        wb.save(file)
        viewer.delete(id)
        updtree()
        messagebox.showinfo("Confirm Delete", "Cliente Eliminado")

    else:
        return True

def updtree():
    global file
    global cont
    wb = load_workbook(filename=file, read_only=TRUE)
    ws = wb1.active
    #max_row_num = ws.max_row
    cont = 1
    for i in range(2, 1):
        cell_obj= ws.cell(row=i, column=1)
        cell_obj1 = ws.cell(row=i, column=2)
        cell_obj2 = ws.cell(row=i, column=3)
        cell_obj3 = ws.cell(row=i, column=4)
        cell_obj4 = ws.cell(row=i, column=5)
        viewer.insert('', END, values=(cont, cell_obj.value, cell_obj1.value, cell_obj2.value, cell_obj3.value, cell_obj4.value))
        cont +=1

def calendar():
    global cal
    global top
    top= Toplevel()
    top.title("Calendar")
    top.resizable(width=FALSE, height=FALSE)
    cal= Calendar(top, selectmode="day", year=2020, month=5, day=22)
    cal.pack(pady=10, fill="both", expand=FALSE)
    okbtn=Button(top, text="Seleccionar Fecha", command=getdate)
    okbtn.pack(pady=5)
    date.set(cal.get_date())

#date= cal.get_date()
def getdate():
    date_entry.config(date.set(cal.get_date()))
    top.destroy()

#Function to edit/update client data (row)
def upd():
    global file
    row_id= int(viewer.focus())
    wb4 = load_workbook(filename=file)
    ws4 = wb4.active
    #updates excel file
    ws4.cell(row=row_id + 1, column=1).value = nmstore_entry.get()
    ws4.cell(row=row_id + 1, column=2).value = actentry.get()
    ws4.cell(row=row_id + 1, column=3).value = status.get()
    ws4.cell(row=row_id + 1, column=4).value = date_entry.get()
    ws4.cell(row=row_id + 1, column=4).value = cmnt_entry.get()
    wb4.save(file)
    messagebox.showinfo("Update Client", "Cliente Actualizado")
    #updates treeview
    sname= store.get()
    oname= name.get()
    stt= status.get()
    dt= date.get()
    cmt= coment.get()
    viewer.item(row_id, values=(row_id,sname, stt, oname, dt,cmt))

    #cleans inputs boxes
    clearinput()

# Menu bar
barMenu = Menu(root)
root.config(menu=barMenu)

#Creación de Menu Bar

# File Menu
file_menu = Menu(barMenu, tearoff=0)
barMenu.add_cascade(label="File", menu=file_menu)
file_menu.add_command(label="New File", command=newfile)
file_menu.add_command(label="Open File", command=openfile)
file_menu.add_command(label="Save")
file_menu.add_separator()    #Separador
file_menu.add_command(label="Close")

#Edit Menu
edit_menu = Menu(barMenu, tearoff=0)
barMenu.add_cascade(label="Edit", menu=edit_menu)
edit_menu.add_command(label="Cut")
edit_menu.add_command(label="Copy")

#Help Menu
hlp_menu = Menu(barMenu, tearoff=0)
barMenu.add_cascade(label="Help", menu=hlp_menu)
hlp_menu.add_command(label="Usa la Fuerza")

#Frames
wrapper1 = LabelFrame(root, text="Customer List")
#wrapper2 = LabelFrame(root, text="Search")
wrapper3 = LabelFrame(root, text="Customer Data")
wrapper1.pack(fill="both",  padx=10, pady=10)
#wrapper2.pack(fill="both", expand="yes", padx=20, pady=10)
wrapper3.pack(fill="both", expand="no", padx=20, pady=10)

#TREEVIEW SCROLL

scroll= Scrollbar(wrapper1)
scroll.pack(side=RIGHT, fill=Y)

#TREE VIEW
viewer = ttk.Treeview(wrapper1, height=6, column=(1,2,3,4,5,6), show='headings', yscrollcommand=scroll.set)
viewer.pack(pady=15)

#configure Scrollbar
scroll.config(command=viewer.yview)


#labels fecha y hora
hoylabel= Label(wrapper1, text="Hoy es: ")
hoylabel.pack(side=LEFT, padx=10, pady=10)
horalabel= Label(wrapper1, text="Hora: ")
horalabel.pack(side=LEFT, padx=10, pady=10)
totallabel= Label(wrapper1, text="Total clientes: ")
totallabel.pack(side=RIGHT, padx=200, pady=5)

clockdate()

viewer.heading(1, text="ID")
viewer.column(1, width=20, anchor=CENTER, minwidth=20)
viewer.heading(2, text="Hotel/Casa de Buceo")
viewer.column(2, width=120, anchor=CENTER, minwidth=100)
viewer.heading(3, text="Fecha")
viewer.column(3, width=50, anchor=CENTER, minwidth=40)
viewer.heading(4, text="Actividad o Paciente")
viewer.column(4, width=120, anchor=CENTER, minwidth=110)
viewer.heading(5, text="Obersvaciones/Servicio")
viewer.column(5, width=130, anchor=CENTER, minwidth=95)
viewer.heading(6, text="Notas")
viewer.column(6, width=100, anchor=CENTER, minwidth=85)
viewer.bind('<Double-1>', getrowsel)

#Input Data Section
nmstore_lbl = Label(wrapper3, text="Hotel/Casa de Buceo")
nmstore_lbl.grid(row=0, column=0, padx=5, pady=20)
nmstore_entry = Entry(wrapper3, textvariable=store)
nmstore_entry.grid(row=0, column=1, padx=5, pady=20)

actlbl = Label(wrapper3, text="Actividad o Paciente")
actlbl.grid(row=1, column=0, padx=5, pady=1)
actentry = Entry(wrapper3, textvariable=name)
actentry.grid(row=1, column=1, padx=10, ipady=20)

fechalbl = Label(wrapper3, text="Fecha", font=("bold", 10))
fechalbl.grid(row=0, column=2, padx=5, pady=20)
fechaentry = Entry(wrapper3, textvariable=status)
fechaentry.grid(row=0, column=3, padx=5, pady=1)

'''
list1 = ['Cerrado', 'En proceso', 'Cancelado']
droplist = OptionMenu(wrapper3, status, *list1)
droplist.config(width=14)
status.set('------------')
droplist.grid(row=0, column=3, padx=3, pady=5)
'''
obslbl=Label(wrapper3, text="Observaciones/Servicio")
obslbl.grid(row=1, column=2, padx=3, pady=5)
date_entry=Entry(wrapper3, textvariable=date)
date_entry.grid(row=1, column=3, padx=10, ipady=20)

noteslbl=Label(wrapper3, text="Notas")
noteslbl.grid(row=0, column=5, padx=10, pady=15)
cmnt_entry=Entry(wrapper3, textvariable=coment)
cmnt_entry.grid(row=1, column=5, padx=10, ipady=20)

#Buttons
add_btn=Button(wrapper3, text="Add New", bg='green', fg='white', command=input)
up_btn=Button(wrapper3, text="Update", bg='blue', fg='white', command=upd)
delete_btn=Button(wrapper3, text="Delete", bg='red', fg='white', command=delete)
#show_btn=Button(wrapper3, text="Show Calendar", command=calendar)
#show_btn.grid(row=3, column=3, pady=5)
add_btn.grid(row=4, column=1, pady=15)
up_btn.grid(row=4, column=2, pady=15)
delete_btn.grid(row=4, column=3, pady=15)

#status bar
statbar= Label(wrapper3, text="No File:", bd=1, pady=5, anchor=W)
statbar.grid(row=5, column=0, columnspan=4, sticky=W)

#END OF FRONT END

#sby.configure(command=viewer.yview)
#viewer.configure(yscrollcommand=sby.set)

root.mainloop()